"""
Route: he_thong_route.py — Quản lý hệ thống (chỉ ADMIN)
GET  /he-thong/nguoi-dung              → danh sách người dùng
POST /he-thong/nguoi-dung/them         → tạo mới người dùng
POST /he-thong/nguoi-dung/{id}/sua     → sửa thông tin người dùng
POST /he-thong/nguoi-dung/{id}/doi-mk  → đặt lại mật khẩu
POST /he-thong/nguoi-dung/{id}/xoa     → vô hiệu hóa / khôi phục
POST /he-thong/nguoi-dung/{id}/duyet   → duyệt tài khoản chờ duyệt
"""

import io
import os, sys
from datetime import date
from typing import Optional
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from config.database import get_db
from api.routes._auth_helper import yeu_cau_quyen_admin
import services.nguoi_dung_service as nd_service
import services.nhat_ky_service as nhat_ky_service
import repositories.don_vi_repository as dv_repo

router = APIRouter()
templates = Jinja2Templates(directory=os.path.join(_ROOT, "templates"))


def _to_int(val: Optional[str]) -> Optional[int]:
    try:
        return int(val) if val and str(val).strip() else None
    except (ValueError, TypeError):
        return None


def _err(url: str, msg: str) -> RedirectResponse:
    return RedirectResponse(url=f"{url}?loi={quote(msg)}", status_code=302)


# ── DANH SÁCH ─────────────────────────────────────────────────────────────

@router.get("/nguoi-dung", response_class=HTMLResponse)
async def danh_sach(
    request: Request,
    loi: Optional[str] = None,
    ok: Optional[str] = None,
    user=Depends(yeu_cau_quyen_admin),
    conn=Depends(get_db),
):
    nguoi_dung_list = nd_service.lay_tat_ca(conn)
    don_vi_list     = dv_repo.lay_tat_ca(conn)
    map_don_vi      = {d.id: d for d in don_vi_list}
    return templates.TemplateResponse("he_thong/nguoi_dung.html", {
        "request":          request,
        "user":             user,
        "nguoi_dung_list":  nguoi_dung_list,
        "don_vi_list":      don_vi_list,
        "map_don_vi":       map_don_vi,
        "loi":              loi,
        "ok":               ok,
    })


# ── TẠO MỚI ───────────────────────────────────────────────────────────────

@router.post("/nguoi-dung/them")
async def them(
    ten_dang_nhap: str = Form(...),
    mat_khau:      str = Form(...),
    ho_ten:        str = Form(...),
    loai_quyen:    str = Form("XEM"),
    chuc_vu:       Optional[str] = Form(None),
    don_vi_id:     Optional[str] = Form(None),
    so_dien_thoai: Optional[str] = Form(None),
    email:         Optional[str] = Form(None),
    user=Depends(yeu_cau_quyen_admin),
    conn=Depends(get_db),
):
    try:
        nd_service.tao_boi_admin(
            conn, ten_dang_nhap, mat_khau, ho_ten, loai_quyen,
            chuc_vu=chuc_vu or None,
            don_vi_id=_to_int(don_vi_id),
            so_dien_thoai=so_dien_thoai or None,
            email=email or None,
            admin_id=user["id"],
        )
    except nd_service.NguoiDungServiceError as e:
        return _err("/he-thong/nguoi-dung", str(e))
    return RedirectResponse(url="/he-thong/nguoi-dung?ok=1", status_code=302)


# ── SỬA THÔNG TIN ─────────────────────────────────────────────────────────

@router.post("/nguoi-dung/{id}/sua")
async def sua(
    id: int,
    ho_ten:        str = Form(...),
    loai_quyen:    str = Form(...),
    chuc_vu:       Optional[str] = Form(None),
    don_vi_id:     Optional[str] = Form(None),
    so_dien_thoai: Optional[str] = Form(None),
    email:         Optional[str] = Form(None),
    is_active:     Optional[str] = Form("1"),
    user=Depends(yeu_cau_quyen_admin),
    conn=Depends(get_db),
):
    try:
        nd_service.sua_thong_tin(
            conn, id, ho_ten, loai_quyen,
            chuc_vu=chuc_vu or None,
            don_vi_id=_to_int(don_vi_id),
            so_dien_thoai=so_dien_thoai or None,
            email=email or None,
            is_active=1 if is_active == "1" else 0,
        )
    except nd_service.NguoiDungServiceError as e:
        return _err("/he-thong/nguoi-dung", str(e))
    return RedirectResponse(url="/he-thong/nguoi-dung?ok=1", status_code=302)


# ── ĐỔI MẬT KHẨU (admin đặt lại) ────────────────────────────────────────

@router.post("/nguoi-dung/{id}/doi-mk")
async def doi_mat_khau(
    id: int,
    mat_khau_moi: str = Form(...),
    user=Depends(yeu_cau_quyen_admin),
    conn=Depends(get_db),
):
    try:
        nd_service.doi_mat_khau_admin(conn, id, mat_khau_moi)
    except nd_service.NguoiDungServiceError as e:
        return _err("/he-thong/nguoi-dung", str(e))
    return RedirectResponse(url="/he-thong/nguoi-dung?ok=1", status_code=302)


# ── VÔ HIỆU / KHÔI PHỤC ──────────────────────────────────────────────────

@router.post("/nguoi-dung/{id}/xoa")
async def xoa(
    id: int,
    user=Depends(yeu_cau_quyen_admin),
    conn=Depends(get_db),
):
    # Ngăn admin tự vô hiệu chính mình
    if id == user["id"]:
        return _err("/he-thong/nguoi-dung", "Không thể vô hiệu hóa tài khoản đang đăng nhập.")
    try:
        nd = nd_service.lay_theo_id(conn, id)
        if nd.is_active:
            nd_service.vo_hieu_hoa(conn, id)
        else:
            nd_service.khoi_phuc(conn, id)
    except nd_service.NguoiDungServiceError as e:
        return _err("/he-thong/nguoi-dung", str(e))
    return RedirectResponse(url="/he-thong/nguoi-dung?ok=1", status_code=302)


# ── DUYỆT TÀI KHOẢN ──────────────────────────────────────────────────────

# ── NHẬT KÝ HỆ THỐNG ──────────────────────────────────────────────────────

@router.get("/nhat-ky", response_class=HTMLResponse)
async def xem_nhat_ky(
    request: Request,
    tab: str = "hoat_dong",
    user=Depends(yeu_cau_quyen_admin),
    conn=Depends(get_db),
):
    nhat_ky = nhat_ky_service.lay_nhat_ky(conn, limit=300)
    dang_nhap_log = nhat_ky_service.lay_dang_nhap_log(conn, limit=300)
    return templates.TemplateResponse("he_thong/nhat_ky.html", {
        "request": request,
        "user": user,
        "nhat_ky": nhat_ky,
        "dang_nhap_log": dang_nhap_log,
        "tab": tab,
    })

# ── XUẤT DỮ LIỆU RA EXCEL ─────────────────────────────────────────────────

# Bảng xuất theo thứ tự phụ thuộc FK, đúng với thứ tự seed
_TABLES = [
    ("cap_quan_ly",      "SELECT id, ma_cap, ten_cap, mo_ta, thu_tu_hien_thi, is_active FROM cap_quan_ly"),
    ("cap_duong",        "SELECT id, ma_cap, ten_cap, mo_ta, thu_tu_hien_thi, is_active FROM cap_duong"),
    ("ket_cau_mat_duong","SELECT id, ma_ket_cau, ten_ket_cau, mo_ta, thu_tu_hien_thi, is_active FROM ket_cau_mat_duong"),
    ("tinh_trang",       "SELECT id, ma_tinh_trang, ten_tinh_trang, mo_ta, mau_hien_thi, thu_tu_hien_thi, is_active FROM tinh_trang"),
    ("don_vi",           "SELECT id, ma_don_vi, ten_don_vi, ten_viet_tat, cap_don_vi, parent_id, dia_chi, so_dien_thoai, email, is_active FROM don_vi"),
    ("nguoi_dung",       "SELECT id, ten_dang_nhap, ho_ten, chuc_vu, don_vi_id, so_dien_thoai, email, loai_quyen, is_active, is_approved, created_at FROM nguoi_dung"),
    ("tuyen_duong",      "SELECT id, ma_tuyen, ten_tuyen, cap_quan_ly_id, don_vi_quan_ly_id, diem_dau, diem_cuoi, chieu_dai_quan_ly, chieu_dai_thuc_te, nam_xay_dung, nam_hoan_thanh, ghi_chu FROM tuyen_duong"),
    ("doan_tuyen",       "SELECT id, ma_doan, tuyen_id, cap_duong_id, tinh_trang_id, ket_cau_mat_id, ly_trinh_dau, ly_trinh_cuoi, chieu_dai_thuc_te, chieu_rong_mat_min, chieu_rong_mat_max, chieu_rong_nen_min, chieu_rong_nen_max, don_vi_bao_duong_id, nam_lam_moi, ngay_cap_nhat_tinh_trang, ghi_chu FROM doan_tuyen"),
    ("doan_di_chung",    "SELECT id, ma_doan_di_chung, tuyen_di_chung_id, tuyen_chinh_id, doan_id, ly_trinh_dau_di_chung, ly_trinh_cuoi_di_chung, ly_trinh_dau_tuyen_chinh, ly_trinh_cuoi_tuyen_chinh, ghi_chu FROM doan_di_chung"),
]

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


@router.get("/xuat-du-lieu")
async def xuat_du_lieu(request: Request, user=Depends(yeu_cau_quyen_admin), conn=Depends(get_db)):
    """Xuất toàn bộ DB ra file Excel — mỗi bảng một sheet."""
    nhat_ky_service.ghi_hoat_dong(
        conn,
        hanh_dong="XUẤT EXCEL",
        nguoi_dung_id=user.get("id"),
        ho_ten=user.get("ho_ten"),
        doi_tuong="Toàn bộ dữ liệu",
        mo_ta="Xuất toàn bộ DB ra file Excel",
        ip_address=request.client.host if request.client else None,
    )
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Xóa sheet trắng mặc định

    for sheet_name, sql in _TABLES:
        rows = conn.execute(sql).fetchall()
        ws = wb.create_sheet(title=sheet_name)

        if not rows:
            ws.append([f"(Bảng {sheet_name} chưa có dữ liệu)"])
            continue

        # Header
        headers = list(rows[0].keys())
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill  = _HEADER_FILL
            cell.font  = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
        ws.row_dimensions[1].height = 22

        # Dữ liệu
        for row in rows:
            ws.append(list(row))

        # Tự động điều chỉnh độ rộng cột
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"giao_thong_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/nguoi-dung/{id}/duyet")
async def duyet(
    id: int,
    loai_quyen: str = Form("XEM"),
    user=Depends(yeu_cau_quyen_admin),
    conn=Depends(get_db),
):
    try:
        nd_service.duyet_tai_khoan(conn, id, approved_by_id=user["id"],
                                   loai_quyen=loai_quyen)
    except nd_service.NguoiDungServiceError as e:
        return _err("/he-thong/nguoi-dung", str(e))
    return RedirectResponse(url="/he-thong/nguoi-dung?ok=1", status_code=302)
