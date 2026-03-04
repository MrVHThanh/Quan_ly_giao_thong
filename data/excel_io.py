"""
excel_io.py — Xuất / Import dữ liệu giao thông từ/ra file Excel.

Cấu trúc file Excel (3 sheet):
  - TuyenDuong      ← tuyen_duong_data.py  (TUYEN_CONFIG)
  - DoanTuyen       ← doan_tuyen_data.py   (DOAN_CONFIG)
  - DoanDiChung     ← doan_di_chung_data.py (DOAN_DI_CHUNG_CONFIG)

Cách dùng:
  python excel_io.py export          # Xuất dữ liệu từ các file .py → Excel
  python excel_io.py import          # Import dữ liệu từ Excel → in ra màn hình
  python excel_io.py import --save   # Import và ghi lại vào các file .py
"""

import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# CẤU HÌNH
# =====================================================================

OUTPUT_FILE = "giao_thong_data.xlsx"

# Màu header
COLOR_HEADER_QL   = "1F4E79"   # xanh đậm — Tuyến đường
COLOR_HEADER_DOAN = "375623"   # xanh lá đậm — Đoạn tuyến
COLOR_HEADER_DDC  = "7B3F00"   # nâu — Đoạn đi chung

# =====================================================================
# CỘT ĐỊNH NGHĨA
# =====================================================================

TUYEN_COLS = [
    ("ma_tuyen",          "Mã tuyến",           12),
    ("ten_tuyen",         "Tên tuyến",           30),
    ("ma_cap_quan_ly",    "Cấp quản lý",         14),
    ("chieu_dai_quan_ly", "CD quản lý (km)",     16),
    ("chieu_dai_thuc_te", "CD thực tế (km)",     16),
    ("ma_don_vi_quan_ly", "Đơn vị quản lý",      18),
    ("diem_dau",          "Điểm đầu",            35),
    ("diem_cuoi",         "Điểm cuối",           35),
    ("lat_dau",           "Lat đầu",             12),
    ("lng_dau",           "Lng đầu",             12),
    ("lat_cuoi",          "Lat cuối",            12),
    ("lng_cuoi",          "Lng cuối",            12),
    ("nam_xay_dung",      "Năm xây dựng",        14),
    ("nam_hoan_thanh",    "Năm hoàn thành",      16),
    ("ghi_chu",           "Ghi chú",             40),
]

DOAN_COLS = [
    ("ma_doan",             "Mã đoạn",             14),
    ("ma_tuyen",            "Mã tuyến",            12),
    ("ma_cap_duong",        "Cấp đường",           12),
    ("ly_trinh_dau",        "Lý trình đầu (km)",   18),
    ("ly_trinh_cuoi",       "Lý trình cuối (km)",  18),
    ("chieu_dai_thuc_te",   "CD thực tế (km)",     16),
    ("chieu_rong_mat_min",  "Rộng mặt min (m)",    16),
    ("chieu_rong_mat_max",  "Rộng mặt max (m)",    16),
    ("chieu_rong_nen_min",  "Rộng nền min (m)",    16),
    ("chieu_rong_nen_max",  "Rộng nền max (m)",    16),
    ("ma_don_vi_bao_duong", "Đơn vị bảo dưỡng",   18),
    ("ghi_chu",             "Ghi chú",             50),
]

DDC_COLS = [
    ("ma_tuyen_di_chung", "Tuyến đi chung",      16),
    ("ma_doan",           "Mã đoạn (tuyến chủ)", 22),
    ("ly_trinh_dau",      "Lý trình đầu (km)",   18),
    ("ly_trinh_cuoi",     "Lý trình cuối (km)",  18),
    ("ghi_chu",           "Ghi chú",             55),
]

# =====================================================================
# HELPER
# =====================================================================

def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_sheet(ws, col_defs, data_rows, header_color):
    """Ghi header + dữ liệu vào một worksheet."""

    # --- Header ---
    keys   = [c[0] for c in col_defs]
    labels = [c[1] for c in col_defs]
    widths = [c[2] for c in col_defs]

    for col_idx, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # --- Dữ liệu ---
    for row_idx, record in enumerate(data_rows, start=2):
        fill_color = "F2F2F2" if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, key in enumerate(keys, start=1):
            val  = record.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=fill_color)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = _thin_border()

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _read_sheet(ws, col_defs):
    """Đọc dữ liệu từ worksheet, trả về list[dict]."""
    keys = [c[0] for c in col_defs]
    result = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        record = {}
        for i, key in enumerate(keys):
            val = row[i] if i < len(row) else None
            # Ép kiểu số nguyên về int nếu là năm
            if key in ("nam_xay_dung", "nam_hoan_thanh") and val is not None:
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            record[key] = val if val != "" else None
        result.append(record)

    return result

# =====================================================================
# IMPORT DỮ LIỆU TỪ CÁC FILE .py
# =====================================================================

def _load_from_py():
    """Nạp dữ liệu từ 3 file data .py."""
    import importlib.util, pathlib

    def load_module(path, name):
        spec = importlib.util.spec_from_file_location(name, path)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod

    base = pathlib.Path(__file__).parent

    tuyen_mod = load_module(base / "tuyen_duong_data.py", "tuyen_duong_data")
    doan_mod  = load_module(base / "doan_tuyen_data.py",  "doan_tuyen_data")
    ddc_mod   = load_module(base / "doan_di_chung_data.py", "doan_di_chung_data")

    return (
        tuyen_mod.TUYEN_CONFIG,
        doan_mod.DOAN_CONFIG,
        ddc_mod.DOAN_DI_CHUNG_CONFIG,
    )

# =====================================================================
# XUẤT EXCEL
# =====================================================================

def export_excel(output_path=OUTPUT_FILE):
    tuyen_data, doan_data, ddc_data = _load_from_py()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # xoá sheet mặc định

    ws_tuyen = wb.create_sheet("TuyenDuong")
    ws_doan  = wb.create_sheet("DoanTuyen")
    ws_ddc   = wb.create_sheet("DoanDiChung")

    _write_sheet(ws_tuyen, TUYEN_COLS, tuyen_data, COLOR_HEADER_QL)
    _write_sheet(ws_doan,  DOAN_COLS,  doan_data,  COLOR_HEADER_DOAN)
    _write_sheet(ws_ddc,   DDC_COLS,   ddc_data,   COLOR_HEADER_DDC)

    wb.save(output_path)
    print(f"✅ Xuất thành công: {output_path}")
    print(f"   TuyenDuong  : {len(tuyen_data)} bản ghi")
    print(f"   DoanTuyen   : {len(doan_data)} bản ghi")
    print(f"   DoanDiChung : {len(ddc_data)} bản ghi")

# =====================================================================
# IMPORT EXCEL
# =====================================================================

def import_excel(input_path=OUTPUT_FILE, save_to_py=False):
    if not os.path.exists(input_path):
        print(f"❌ Không tìm thấy file: {input_path}")
        return None, None, None

    wb = openpyxl.load_workbook(input_path)

    tuyen_data = _read_sheet(wb["TuyenDuong"],  TUYEN_COLS)
    doan_data  = _read_sheet(wb["DoanTuyen"],   DOAN_COLS)
    ddc_data   = _read_sheet(wb["DoanDiChung"], DDC_COLS)

    print(f"✅ Import thành công: {input_path}")
    print(f"   TuyenDuong  : {len(tuyen_data)} bản ghi")
    print(f"   DoanTuyen   : {len(doan_data)} bản ghi")
    print(f"   DoanDiChung : {len(ddc_data)} bản ghi")

    if save_to_py:
        _save_to_py(tuyen_data, doan_data, ddc_data)

    return tuyen_data, doan_data, ddc_data


def _save_to_py(tuyen_data, doan_data, ddc_data):
    """Ghi lại dữ liệu đã import vào các file .py tương ứng."""
    import pathlib

    base = pathlib.Path(__file__).parent

    def _write_config(path, var_name, data):
        lines = [f'"""\nAuto-generated by excel_io.py. Chỉnh sửa trực tiếp hoặc qua Excel.\n"""\n\n']
        lines.append(f"{var_name} = [\n")
        for record in data:
            lines.append("    {\n")
            for k, v in record.items():
                lines.append(f"        {k!r}: {v!r},\n")
            lines.append("    },\n")
        lines.append("]\n")
        path.write_text("".join(lines), encoding="utf-8")
        print(f"   Đã ghi: {path}")

    _write_config(base / "tuyen_duong_data.py",    "TUYEN_CONFIG",          tuyen_data)
    _write_config(base / "doan_tuyen_data.py",     "DOAN_CONFIG",           doan_data)
    _write_config(base / "doan_di_chung_data.py",  "DOAN_DI_CHUNG_CONFIG",  ddc_data)
    print("✅ Đã ghi lại các file .py")

# =====================================================================
# CLI
# =====================================================================

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "export"

    if cmd == "export":
        export_excel()

    elif cmd == "import":
        save = "--save" in sys.argv
        import_excel(save_to_py=save)

    else:
        print(__doc__)